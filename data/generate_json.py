from pathlib import Path

from openpyxl import load_workbook
import json


class DataManager:
    def __init__(self):
        self.workbook = load_workbook('data.xlsx', data_only=True)
        self.data = {
            'categories': {

            },
            'structure': {
                'categories': [
                    'header armor',
                    'helm',
                    'chest',
                    'gauntlets',
                    'legs',

                    'header equipment',
                    'weapon',
                    'shield',
                    'talisman',

                    'header magic',
                    'sorcery',
                    'incantation',

                    'header skills',
                    'ashes',
                    'spirits',

                    'header miscellaneous',
                    'cookbook',
                    'gesture',
                    'crystaltear',
                    'remembrance',
                ],
                'subcategories': {
                    'helm': None,
                    'chest': None,
                    'gauntlets': None,
                    'legs': None,
                    'set': None,
                    'weapon': [
                        'header sword',
                        'weapons_dagger',
                        'weapons_straightsword',
                        'weapons_lightgreatsword',
                        'weapons_greatsword', #dlc
                        'weapons_colossalsword',
                        'weapons_curvedsword',
                        'weapons_curvedgreatsword',
                        'weapons_katana',
                        'weapons_greatkatana', #dlc
                        'weapons_thrustingsword',
                        'weapons_heavythrustingsword',
                        'weapons_backhandblade', #dlc

                        'header sword',
                        'weapons_axe',
                        'weapons_greataxe',
                        'weapons_hammer',
                        'weapons_warhammer',
                        'weapons_spear',
                        'weapons_greatspear',
                        'weapons_halberd',

                        'header ranged',
                        'weapons_lightbow',
                        'weapons_bow',
                        'weapons_greatbow',
                        'weapons_crossbow',
                        'weapons_ballista',

                        'header other',
                        'weapons_colossalweapon',
                        'weapons_twinblade',
                        'weapons_reaper',
                        'weapons_whip',
                        'weapons_flail',
                        'weapons_claw',
                        'weapons_beastclaw',
                        'weapons_fist',
                        'weapons_torch',
                        'weapons_hand-to-handarts', #dlc
                        'weapons_perfumebottles', #dlc
                        'weapons_throwingblades', #dlc

                        'header magic',
                        'weapons_glintstonestaff',
                        'weapons_sacredseal',
                    ],
                    'shield': [
                        'header shield',
                        'shield_smallshield',
                        'shield_mediumshield',
                        'shield_greatshield',
                        'shield_thrustingshield', #dlc
                    ],
                    'talisman': None,
                    'incantation': [
                        'header incantation_order',
                        'incantation_erdtreeincantation',
                        'incantation_goldenorderincantation',
                        'incantation_twofingerincantation',
                        'incantation_miquellanincantation', #dlc
                        'incantation_spiralincantation', #dlc

                        'header incantation_fire',
                        'incantation_firegiantincantation',
                        'incantation_firemonkincantation',
                        'incantation_godskinapostleincantation',
                        'incantation_messmersfireincantation', #dlc

                        'header incantation_chaos',
                        'incantation_bestialincantation',
                        'incantation_bloodincantation',
                        'incantation_frenziedflameincantation',
                        'incantation_servantsofrotincantation',

                        'header incantation_dragon',
                        'incantation_dragoncommunionincantation',
                        'incantation_dragoncultincantation',
                    ],
                    'sorcery': [
                        'header sorcery_type',
                        'sorcery_glintstonesorceries',
                        'sorcery_gravitysorceries',
                        'sorcery_nightsorceries',
                        'sorcery_primevalsorceries',
                        'sorcery_fingersorceries', #dlc

                        'header sorcery_faction',
                        'sorcery_cariansorceries',
                        'sorcery_claymensorceries',
                        'sorcery_crystalliansorceries',
                        'sorcery_fullmoonsorceries',
                        'sorcery_lorettassorceries',

                        'header sorcery_element',
                        'sorcery_magmasorceries',
                        'sorcery_snowwitchsorceries',

                        'header sorcery_chaos',
                        'sorcery_aberrantsorceries',
                        'sorcery_deathsorceries',
                        'sorcery_scadutreesorceries',  # dlc
                    ],
                    'ashes': None,
                    'spirits': None,
                    'cookbook': None,
                    'gesture': None,
                    'crystaltear': None,
                    'remembrance': None
                }
            },
            'upgrades': []
        }

    def read_sheet(self, sheet_name):
        def generate_template():
            return {
                'id': None,
                'dlc': False,
                'icon': None,
                'wiki': None,
                'position': 0,
                'collected': False,
                'name': {},
                'category': {},
                'subcategory': {}
            }

        def cell_value(cell):
            value = cell.value
            if value == 'none':
                return None
            if value == 'false':
                return False
            if value == 'true':
                return True
            return value

        sheet = self.workbook[sheet_name]
        category = {
            'id': sheet[3][6].value,
            'en': {
                'singular': sheet[3][7].value.split(';')[0],
                'plural': sheet[3][7].value.split(';')[1],
            },
            'de': {
                'singular': sheet[3][8].value.split(';')[0],
                'plural': sheet[3][8].value.split(';')[1],
            },
            'stats': {
                'total': 1,
                'total_dlc': 1,
                'total_alternative': 1,
                'total_alternative_dlc': 1,
                'collected': 0,
                'percentage': 0
            },
            'subcategories': {},
            'items': []
            }

        main_game_suffix = '_elden_ring_wiki_guide_200px.png'
        dlc_suffix = '_elden_ring_shadow_of_the_erdtree_dlc_wiki_guide_200px.png'

        for i, row in enumerate(sheet.iter_rows(min_row=4)):
            if cell_value(row[0]) is not None:
                template = generate_template()
                template['id'] = cell_value(row[0])
                if cell_value(row[12]) == 'True':
                    template['dlc'] = True
                    suffix = dlc_suffix
                else:
                    suffix = main_game_suffix
                if cell_value(row[1]) is None:
                    try:
                        if sheet_name == 'sets':
                            template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + '.png'
                        if sheet_name in ['helms', 'chests', 'gauntlets', 'legs']:
                            if template['dlc']:
                                if sheet_name == 'chests':
                                    template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + '_' + category['en']['singular'].lower().replace(' ', '_') + '_armor' + f'{suffix}'
                                else:
                                    template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + '_' + category['en']['singular'].lower().replace(' ', '_') + f'{suffix}'
                            else:
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'{suffix}'
                        if sheet_name == 'weapons':
                            template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'{suffix}'
                            if cell_value(row[9]) == 'weapons_ballista':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_ballista_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_axe':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_bow':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_claw':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_claw_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_colossalsword':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_colossal_swords{suffix}'
                            if cell_value(row[9]) == 'weapons_colossalweapon':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_colossal_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_greataxe':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_greataxe{suffix}'
                            if cell_value(row[9]) == 'weapons_crossbow':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_curvedsword':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_curved_sword_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_curvedgreatsword':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_curved_greatsword_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_dagger':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_dagger_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_fist':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_fist_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_flail':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_glintstonestaff':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_glintstonestaff_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_greatspear':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_greatspear_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_greataxe':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_greataxe_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_halberd':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_halberd_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_hammer':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_hammer_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_katana':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_katana_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_sacredseal':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_sacred_seal_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_spear':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_spear_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_straightsword':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_straight_sword_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_thrustingsword':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_thrusting_sword_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_warhammer':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_warhammer_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_whip':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_torch':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_reaper':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_reaper_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_heavythrustingsword':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_heavy_thrusting_sword_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_greatsword':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_greatbow':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_weapon{suffix}'
                            if cell_value(row[9]) == 'weapons_twinblade':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'_twinblade_weapon{suffix}'

                        if sheet_name == 'ashes':
                            if template['dlc']:
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/ash_of_war_' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + '_' + category['en']['plural'].lower().replace(' ', '_') + f'{suffix}'
                            else:
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/ash_of_war_' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'{suffix}'

                        if sheet_name in ['spirits', 'shields']:
                            if template['dlc'] and sheet_name == 'spirits':
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + '_' + category['en']['singular'].lower().replace(' ', '_') + f'{suffix}'
                            else:
                                template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'{suffix}'
                        if sheet_name in ['incantations', 'sorceries', 'talismans', 'cookbooks', 'crystaltears', 'remembrances']:
                            if template['dlc']:
                                if cell_value(row[4]).lower() in [
                                    'spira',
                                    'giant golden arc',
                                    'golden arc',
                                ]:
                                    template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + '_' + 'sorceries'.lower() + f'{suffix}'
                                elif sheet_name in ['talismans', 'crystaltears', 'incantations']:
                                    template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + '_' + category['en']['singular'].lower().replace(' ', '_') + f'{suffix}'
                                elif sheet_name in ['cookbooks']:
                                    template['icon'] = ('https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "").replace('1', '').replace('2', '').replace('3', '').replace('4', '').replace('5', '').replace('6', '').replace('7', '').replace('8', '').replace('9', '').replace('0', '').replace('[', '').replace(']', '') + category['en']['plural'].lower() + f'{suffix}')
                                else:
                                    template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + '_' + category['en']['plural'].lower() + f'{suffix}'
                            else:
                                if sheet_name == 'remembrances':
                                    template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'{suffix}'
                                elif sheet_name in ['crystaltears']:
                                    template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "") + f'{suffix}'
                                else:
                                    template['icon'] = 'https://eldenring.wiki.fextralife.com/file/Elden-Ring/' + cell_value(row[4]).lower().replace(' ', '_').replace("'", "").replace(',', '').replace('!', '').replace(':', '') + '_' + category['en']['singular'].lower().replace(' ', '_') + f'{suffix}'
                        if sheet_name == 'talismans':
                            template['icon'] = template['icon'].replace('+', '')
                    except Exception as e:
                        print(e)
                else:
                    template['icon'] = cell_value(row[1])

                if cell_value(row[2]) is None:
                    try:
                        if sheet_name in ['sets', 'helms', 'chests', 'gauntlets', 'legs']:
                            template['wiki'] = 'https://eldenring.wiki.fextralife.com/' + cell_value(row[4]).replace(' ', '+')
                        if sheet_name == 'ashes':
                            template['wiki'] = 'https://eldenring.wiki.fextralife.com/' + 'Ash+of+War:+' + cell_value(row[4]).replace(' ', '+')
                        if sheet_name in ['incantations', 'sorceries', 'talismans', 'cookbooks', 'spirits', 'shields', 'weapons', 'crystaltears', 'remembrances']:
                            template['wiki'] = 'https://eldenring.wiki.fextralife.com/' + cell_value(row[4]).replace(' ', '+')
                    except Exception as e:
                        print(e)
                else:
                    template['wiki'] = cell_value(row[2])
                if sheet_name == 'cookbooks':
                    template['wiki'] = template['wiki'].replace('[', '(').replace(']', ')')

                template['position'] = cell_value(row[3])
                template['name'] = {
                    'en': cell_value(row[4]),
                    'de': cell_value(row[5])
                }
                template['category'] = {
                    'id': category['id'],
                    'en': cell_value(row[7]),
                    'de': cell_value(row[8])
                }
                template['subcategory'] = {
                    'id': cell_value(row[9]),
                    'en': cell_value(row[10]),
                    'de': cell_value(row[11])
                }
                if sheet_name == 'talismans':
                    template['upgraded'] = False
                    if cell_value(row[13]) is not None:
                        template['upgraded'] = True
                if sheet_name in ['helms', 'chests', 'gauntlets', 'legs']:
                    template['set'] = cell_value(row[13])
                    template['altered'] = False
                    if cell_value(row[14]) is not None:
                        template['altered'] = True

                if category['subcategories'].get(template['subcategory']['id']) is None and template['subcategory']['id'] is not None:
                    category['subcategories'][template['subcategory']['id']] = template['subcategory']
                category['items'].append(template)

        self.data['categories'][category['id']] = category

        category['stats']['total'] = len([[].append(item) for item in self.data['categories'][category['id']]['items'] if not item['dlc']])
        category['stats']['total_dlc'] = len(self.data['categories'][category['id']]['items'])
        if sheet_name == 'talismans':
            category['stats']['total'] = len([[].append(item) for item in self.data['categories'][category['id']]['items'] if not item['dlc'] and not item['upgraded']])
            category['stats']['total_dlc'] = len([[].append(item) for item in self.data['categories'][category['id']]['items'] if not item['upgraded']])
            category['stats']['total_alternative'] = len([[].append(item) for item in self.data['categories'][category['id']]['items'] if not item['dlc']])
            category['stats']['total_alternative_dlc'] = len(self.data['categories'][category['id']]['items'])
        if sheet_name in ['helms', 'chests']:
            category['stats']['total'] = len([[].append(item) for item in self.data['categories'][category['id']]['items'] if not item['dlc'] and not item['altered']])
            category['stats']['total_dlc'] = len([[].append(item) for item in self.data['categories'][category['id']]['items'] if not item['altered']])
            category['stats']['total_alternative'] = len([[].append(item) for item in self.data['categories'][category['id']]['items'] if not item['dlc']])
            category['stats']['total_alternative_dlc'] = len(self.data['categories'][category['id']]['items'])

        for sc in self.data['categories'][category['id']]['subcategories']:
            print(sc)
        print()

    def get_upgrades(self):
        sheet = self.workbook['upgrades']
        upgrades = []

        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            if row[0].value is not None:
                upgrades.append({
                    'new': row[0].value,
                    'old': row[1].value,
                    'category': row[2].value,
                    'version': row[3].value
                })
        self.data['upgrades'] = upgrades

    def run(self):
        self.get_upgrades()
        self.read_sheet('sets')
        self.read_sheet('helms')
        self.read_sheet('chests')
        self.read_sheet('gauntlets')
        self.read_sheet('legs')
        self.read_sheet('weapons')
        self.read_sheet('shields')
        self.read_sheet('talismans')
        self.read_sheet('sorceries')
        self.read_sheet('incantations')
        self.read_sheet('ashes')
        self.read_sheet('spirits')
        self.read_sheet('cookbooks')
        self.read_sheet('gestures')
        self.read_sheet('crystaltears')
        self.read_sheet('remembrances')

        path = Path(__file__).parent.parent / 'static/data.json'
        with open(path, 'w') as file:
            file.write(json.dumps(self.data))
            file.close()

        exit()


DM = DataManager()
DM.run()
